import os
from datetime import datetime
from flask import Flask, render_template, request, redirect, flash, send_file, url_for
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)
app.secret_key = "asddjajs kkdbw"
app.config['SESSION_TYPE'] = 'filesystem'

month_format = {
    12: "01_jan",
    1: "02_feb",
    2: "03_mar",
    3: "04_apr",
    4: "05_may",
    5: "06_jul",
    6: "07_aug",
    7: "08_aug",
    8: "09_sep",
    9: "10_oct",
    10: "11_nov",
    11: "12_dec"
}

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        flash('No file part', 'error')
        return redirect(request.url)

    file = request.files['file']

    if file.filename == '':
        flash('No selected file', 'error')
        return redirect(request.url)

    try:
        filename = file.filename
        file.save(filename)

        # Process the uploaded file
        update_file(filename)
        
        return redirect(url_for('index'))

    except Exception as e:
        flash(f'Error processing the file: {str(e)}', 'error')
        os.remove(filename)  # Delete the file in case of error
        return redirect(request.url)

def update_file(filename):
    try:
        wb = load_workbook(filename=filename)

        current_month = filename[-8:-5]
        current_year = datetime.now().year

        months_to_next = {
            "jan": "February",
            "feb": "March",
            "mar": "April",
            "apr": "May",
            "may": "June",
            "jun": "July",
            "jul": "August",
            "aug": "September",
            "sep": "October",
            "oct": "November",
            "nov": "December",
            "dec": "January"
        }

        forecast = wb["Forecast"]
        actuals = wb["Actuals"]

        if months_to_next[current_month] == "January":
            current_year += 1

        next_month = months_to_next[current_month][:3] + " " + str(current_year)

        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

        forecast_col_array = []
        for column_cell in forecast.iter_cols(1, forecast.max_column):  
            if next_month in column_cell[0].value:
                for data in column_cell[1:]:
                    forecast_col_array.append(data.value)
                    data.value = None
                break

        actuals_col = []
        for column_cell in actuals.iter_cols(1, forecast.max_column):
            if next_month in column_cell[0].value:
                j = 0
                while j < len(forecast_col_array):
                    cell = column_cell[1:][j]
                    cell.value = forecast_col_array[j]
                    cell.fill = green_fill  # Highlight the cell
                    j += 1
                break

        new_file_name = filename.replace(current_month, months_to_next[current_month][:3].lower())

        wb.save(filename=new_file_name)

    except Exception as e:
        flash(f'Error updating the file: {str(e)}', 'error')
        os.remove(filename)  # Delete the file in case of error

if __name__ == '__main__':
    app.run(debug=True)
